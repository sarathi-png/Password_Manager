"""CSV / Excel import with automatic format detection.

Supported formats:
  - chrome   (Chrome / Edge): name,url,username,password,note
  - firefox  : url,username,password,httpRealm,formActionOrigin,guid,...
  - bitwarden: folder,favorite,type,name,notes,fields,reprompt,login_uri,...
  - onepassword: Title,Url,Username,Password,Notes,OTPAuth,...
  - generic  : any table with url/username/password-like columns
"""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field

from openpyxl import load_workbook

MAX_CELL = 4096

HEADER_ALIASES: dict[str, list[str]] = {
    "title": ["title", "name", "site name", "label", "entry", "website name", "item name"],
    "url": ["url", "uri", "website", "site", "web site", "login_uri", "address", "link", "domain"],
    "username": ["username", "user name", "user", "login", "login_username", "email", "e-mail", "mail", "account", "userid", "user id", "user_id"],
    "password": ["password", "pass", "pwd", "passwd", "secret", "key", "login_password", "password_value", "pswd"],
    "notes": ["notes", "note", "comments", "comment", "remarks", "extra", "description", "memo", "details"],
}

FORMAT_SIGNATURES: list[tuple[str, set[str]]] = [
    ("bitwarden", {"loginuri", "loginusername", "loginpassword"}),
    ("onepassword", {"title", "url", "username", "password", "otpauth"}),
    ("firefox", {"httprealm", "formactionorigin"}),
    ("chrome", {"name", "url", "username", "password", "note"}),
]


@dataclass
class ImportRow:
    title: str = ""
    url: str = ""
    username: str = ""
    password: str = ""
    notes: str = ""
    category: str = "other"


@dataclass
class ParsedImport:
    detected_format: str
    headers: list[str]
    rows: list[ImportRow] = field(default_factory=list)
    mapping: dict[str, str] = field(default_factory=dict)


def _norm_header(h: str) -> str:
    return re.sub(r"[^a-z0-9]", "", h.strip().lower())


def _find_column(headers: list[str], field: str) -> str | None:
    for alias in HEADER_ALIASES[field]:
        target = _norm_header(alias)
        for h in headers:
            if _norm_header(h) == target:
                return h
    return None


def detect_format(headers: list[str]) -> str:
    norm = {_norm_header(h) for h in headers}
    for fmt, sig in FORMAT_SIGNATURES:
        if sig.issubset(norm):
            return fmt
    return "generic"


def build_mapping(headers: list[str]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for field in ("title", "url", "username", "password", "notes"):
        col = _find_column(headers, field)
        if col is not None:
            mapping[field] = col
    if "password" not in mapping:
        raise ValueError("No password column detected in the file")
    if "title" not in mapping:
        mapping["title"] = "url" if "url" in mapping else "username" if "username" in mapping else "password"
    return mapping


def _clean(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return text[:MAX_CELL]


def _classify(url: str, title: str) -> str:
    haystack = f"{url} {title}".lower()
    rules = [
        ("email", ["mail.google", "gmail", "outlook", "yahoo", "protonmail", "zoho"]),
        ("banking", ["bank", "paypal", "stripe", "revolut", "wise", "n26", "citi", "chase", "hsbc", "sbi", "icici", "hdfc"]),
        ("social", ["facebook", "twitter", "x.com", "instagram", "linkedin", "reddit", "tiktok", "whatsapp", "telegram", "discord", "snapchat", "youtube"]),
        ("shopping", ["amazon", "flipkart", "ebay", "walmart", "alibaba", "shopify", "etsy", "myntra", "bestbuy"]),
        ("work", ["github", "gitlab", "jira", "slack", "notion", "atlassian", "aws", "azure", "gcp", "google cloud", "digitalocean", "heroku"]),
        ("entertainment", ["netflix", "spotify", "prime", "hulu", "disney", "hotstar", "sony", "crunchyroll", "steam", "epic games"]),
    ]
    for category, keywords in rules:
        for k in keywords:
            if "." in k:
                if re.search(rf"(^|[^a-z0-9.]){re.escape(k)}($|[^a-z0-9.])", haystack):
                    return category
            elif k in haystack:
                return category
    return "other"


def _to_row(record: dict[str, str], mapping: dict[str, str]) -> ImportRow:
    get = lambda field: _clean(record.get(mapping.get(field, ""), ""))
    row = ImportRow(
        title=get("title"),
        url=get("url"),
        username=get("username"),
        password=get("password"),
        notes=get("notes"),
    )
    if not row.title:
        row.title = row.url.split("//")[-1].split("/")[0] if row.url else row.username or "Untitled"
    row.category = _classify(row.url, row.title)
    return row


def parse_csv_bytes(data: bytes) -> ParsedImport:
    text = _decode(data)
    reader = csv.reader(io.StringIO(text))
    try:
        headers = [h.strip() for h in next(reader)]
    except StopIteration:
        raise ValueError("File is empty")

    fmt = detect_format(headers)
    mapping = build_mapping(headers)
    rows: list[ImportRow] = []
    for raw in reader:
        if not raw or all(not cell.strip() for cell in raw):
            continue
        record = dict(zip(headers, raw))
        if not record.get(mapping.get("password", ""), "").strip():
            continue
        rows.append(_to_row(record, mapping))
    return ParsedImport(detected_format=fmt, headers=headers, rows=rows, mapping=mapping)


def parse_xlsx_bytes(data: bytes) -> ParsedImport:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    iterator = ws.iter_rows(values_only=True)
    try:
        headers = [_clean(h) for h in next(iterator)]
    except StopIteration:
        raise ValueError("File is empty")

    fmt = detect_format(headers)
    mapping = build_mapping(headers)
    rows: list[ImportRow] = []
    for values in iterator:
        record = dict(zip(headers, [_clean(v) for v in values]))
        if not record.get(mapping.get("password", ""), "").strip():
            continue
        rows.append(_to_row(record, mapping))
    wb.close()
    return ParsedImport(detected_format=fmt, headers=headers, rows=rows, mapping=mapping)


def parse_import(data: bytes, filename: str) -> ParsedImport:
    if filename.lower().endswith((".xlsx", ".xlsm")):
        return parse_xlsx_bytes(data)
    if filename.lower().endswith(".xls"):
        raise ValueError("Legacy .xls files are not supported; please save as .xlsx or .csv")
    return parse_csv_bytes(data)


def _decode(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")